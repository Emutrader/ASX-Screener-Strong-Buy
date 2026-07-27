#!/usr/bin/env python3
"""
ASX Strong Buy Stock Screener
==============================
Pulls quantitative + analyst-sentiment data for ASX-listed stocks and
produces an Excel workbook with two tabs: "ASX 200" and "Penny Stocks".

Data source: Yahoo Finance, via the `yfinance` library (free, no API key).

IMPORTANT LIMITATIONS (read this before trusting the output):
  - Yahoo Finance's analyst coverage of ASX stocks is inconsistent. Small
    and mid-cap names often have NO analyst coverage at all -> those
    fields will read "No data".
  - "Forecast price Growth next 3/6/12 months" is NOT something Yahoo
    Finance (or any free source) actually publishes broken out by
    horizon. Analysts publish a single ~12-month mean target price.
    This script uses that 12-month target to derive a 12-month growth
    %, then LINEARLY INTERPOLATES 3-month and 6-month figures as simple
    fractions of that. This is a mathematical approximation, not a
    genuine independent 3-month or 6-month analyst forecast. Treat it
    as illustrative only.
  - "Strong Buy Yes/No" is derived from Yahoo's numeric analyst
    recommendation score (1.0 = Strong Buy ... 5.0 = Strong Sell).
    This script marks "Yes" when the mean score is <= 2.0 (i.e. Buy or
    better). It is Yahoo's aggregated view, not a guarantee, not
    financial advice, and not equivalent to any single broker's own
    "Strong Buy" label.
  - "Revenue / Net Profit, 5 years" — free yfinance typically exposes
    only the last 4 annual reporting periods. The script writes
    whatever is available and labels each column with its actual year.
  - The Notes field is the company's own business description
    (Yahoo's `longBusinessSummary`), NOT a forward-looking analyst
    outlook — free sources don't provide that. It's static per company
    and only changes when Yahoo updates its profile data.

This script is a data-gathering and reporting tool. It does not place
trades and is not financial advice.
"""

import sys
import time
import logging
import datetime as dt
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("screener_run.log"),
    ],
)
log = logging.getLogger("asx_screener")

HERE = Path(__file__).resolve().parent
OUTPUT_DIR = HERE / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

HEADERS = [
    "Stock Code",
    "Company Name",
    "Current Price ($)",
    "52 Week High ($)",
    "52 Week Low ($)",
    "Forecast Growth 3mo (%)",
    "Forecast Growth 6mo (%)",
    "Forecast Growth 12mo (%)",
    "Strong Buy (Analysts)",
    "P/E Ratio",
    "P/B Ratio",
    "P/S Ratio",
    "Debt to Equity",
    "Revenue (5yr)",
    "Net Profit (5yr)",
    "Notes - Business Status",
    "Data As Of",
]

RETRY_ATTEMPTS = 3
RETRY_DELAY_SEC = 2


def safe_get(d, key, default=None):
    val = d.get(key, default)
    if val in (None, "", "N/A"):
        return default
    return val


def fmt_money(v):
    return round(v, 3) if isinstance(v, (int, float)) else None


def fmt_pct(v):
    return round(v, 4) if isinstance(v, (int, float)) else None


def fetch_financial_history(ticker_obj, key):
    """Pull up to 5 years of an annual income-statement line item.
    Returns a string like '2025: $1,234M | 2024: $1,100M | ...'
    yfinance free tier typically returns 4 annual periods, not 5."""
    try:
        fin = ticker_obj.financials  # annual income statement, columns = period end dates
        if fin is None or fin.empty or key not in fin.index:
            return "No data"
        row = fin.loc[key].dropna()
        parts = []
        for period_end, value in row.items():
            year = period_end.year if hasattr(period_end, "year") else str(period_end)[:4]
            parts.append(f"{year}: ${value / 1e6:,.1f}M")
        if not parts:
            return "No data"
        return " | ".join(parts)
    except Exception as e:
        log.debug(f"financial history fetch failed for key={key}: {e}")
        return "No data"


def derive_forecast_growth(current_price, target_mean_price):
    """Approximate 3/6/12-month growth from the single analyst 12-month
    mean target price, via straight-line interpolation. See module
    docstring for the caveat on this approximation."""
    if not current_price or not target_mean_price:
        return None, None, None
    total_growth_12mo = (target_mean_price - current_price) / current_price
    growth_3mo = total_growth_12mo * (3 / 12)
    growth_6mo = total_growth_12mo * (6 / 12)
    return fmt_pct(growth_3mo), fmt_pct(growth_6mo), fmt_pct(total_growth_12mo)


def strong_buy_flag(recommendation_mean):
    if recommendation_mean is None:
        return "No data"
    return "Yes" if recommendation_mean <= 2.0 else "No"


def fetch_one(ticker_code):
    """Fetch and shape one row of data for a single ASX ticker (no .AX suffix)."""
    import yfinance as yf  # lazy import: only needed when actually fetching live data
    yf_symbol = f"{ticker_code}.AX"
    last_err = None
    for attempt in range(1, RETRY_ATTEMPTS + 1):
        try:
            t = yf.Ticker(yf_symbol)
            info = t.info or {}

            company_name = safe_get(info, "longName") or safe_get(info, "shortName") or ticker_code
            current_price = safe_get(info, "currentPrice") or safe_get(info, "regularMarketPrice")
            week_high = safe_get(info, "fiftyTwoWeekHigh")
            week_low = safe_get(info, "fiftyTwoWeekLow")
            target_mean = safe_get(info, "targetMeanPrice")
            rec_mean = safe_get(info, "recommendationMean")
            pe = safe_get(info, "trailingPE") or safe_get(info, "forwardPE")
            pb = safe_get(info, "priceToBook")
            ps = safe_get(info, "priceToSalesTrailing12Months")
            de = safe_get(info, "debtToEquity")
            if isinstance(de, (int, float)) and de > 10:
                # Yahoo often reports this as a percentage (e.g. 45.2 meaning 45.2%)
                de = de / 100
            summary = safe_get(info, "longBusinessSummary", "No business summary available.")
            if summary and len(summary) > 400:
                summary = summary[:397] + "..."

            g3, g6, g12 = derive_forecast_growth(current_price, target_mean)

            revenue_hist = fetch_financial_history(t, "Total Revenue")
            profit_hist = fetch_financial_history(t, "Net Income")

            row = {
                "Stock Code": ticker_code,
                "Company Name": company_name,
                "Current Price ($)": fmt_money(current_price),
                "52 Week High ($)": fmt_money(week_high),
                "52 Week Low ($)": fmt_money(week_low),
                "Forecast Growth 3mo (%)": g3,
                "Forecast Growth 6mo (%)": g6,
                "Forecast Growth 12mo (%)": g12,
                "Strong Buy (Analysts)": strong_buy_flag(rec_mean),
                "P/E Ratio": round(pe, 2) if isinstance(pe, (int, float)) else "No data",
                "P/B Ratio": round(pb, 2) if isinstance(pb, (int, float)) else "No data",
                "P/S Ratio": round(ps, 2) if isinstance(ps, (int, float)) else "No data",
                "Debt to Equity": round(de, 2) if isinstance(de, (int, float)) else "No data",
                "Revenue (5yr)": revenue_hist,
                "Net Profit (5yr)": profit_hist,
                "Notes - Business Status": summary,
                "Data As Of": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
            }
            return row
        except Exception as e:
            last_err = e
            log.warning(f"{yf_symbol}: attempt {attempt}/{RETRY_ATTEMPTS} failed ({e})")
            time.sleep(RETRY_DELAY_SEC)

    log.error(f"{yf_symbol}: giving up after {RETRY_ATTEMPTS} attempts ({last_err})")
    return {
        "Stock Code": ticker_code,
        "Company Name": "FETCH FAILED",
        "Current Price ($)": None, "52 Week High ($)": None, "52 Week Low ($)": None,
        "Forecast Growth 3mo (%)": None, "Forecast Growth 6mo (%)": None, "Forecast Growth 12mo (%)": None,
        "Strong Buy (Analysts)": "No data", "P/E Ratio": "No data", "P/B Ratio": "No data",
        "P/S Ratio": "No data", "Debt to Equity": "No data",
        "Revenue (5yr)": "No data", "Net Profit (5yr)": "No data",
        "Notes - Business Status": f"Data fetch failed: {last_err}",
        "Data As Of": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


def load_tickers(csv_path):
    df = pd.read_csv(csv_path)
    return [str(t).strip().upper() for t in df["ticker"].dropna().tolist()]


def fetch_universe(tickers, label):
    rows = []
    total = len(tickers)
    for i, tkr in enumerate(tickers, start=1):
        log.info(f"[{label}] ({i}/{total}) fetching {tkr}...")
        rows.append(fetch_one(tkr))
        time.sleep(0.4)  # be polite to Yahoo's endpoint, avoid rate-limit blocks
    return pd.DataFrame(rows, columns=HEADERS)


# ---------------------------------------------------------------------------
# Excel formatting
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
STRONG_BUY_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
STRONG_BUY_FONT = Font(name="Arial", size=10, color="006100")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

COL_WIDTHS = {
    "Stock Code": 10, "Company Name": 30, "Current Price ($)": 13,
    "52 Week High ($)": 13, "52 Week Low ($)": 13,
    "Forecast Growth 3mo (%)": 15, "Forecast Growth 6mo (%)": 15, "Forecast Growth 12mo (%)": 16,
    "Strong Buy (Analysts)": 15, "P/E Ratio": 10, "P/B Ratio": 10, "P/S Ratio": 10,
    "Debt to Equity": 13, "Revenue (5yr)": 55, "Net Profit (5yr)": 55,
    "Notes - Business Status": 70, "Data As Of": 16,
}

PCT_COLS = {"Forecast Growth 3mo (%)", "Forecast Growth 6mo (%)", "Forecast Growth 12mo (%)"}
MONEY_COLS = {"Current Price ($)", "52 Week High ($)", "52 Week Low ($)"}


def write_sheet(wb, sheet_name, df):
    ws = wb.create_sheet(sheet_name)
    ws.append(HEADERS)
    for c_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for _, data_row in df.iterrows():
        ws.append([data_row[h] for h in HEADERS])

    last_row = ws.max_row
    for r in range(2, last_row + 1):
        is_strong_buy = ws.cell(row=r, column=HEADERS.index("Strong Buy (Analysts)") + 1).value == "Yes"
        for c_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=r, column=c_idx)
            cell.font = STRONG_BUY_FONT if is_strong_buy else BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(header in ("Notes - Business Status", "Revenue (5yr)", "Net Profit (5yr)")))
            if is_strong_buy:
                cell.fill = STRONG_BUY_FILL
            if header in PCT_COLS and isinstance(cell.value, (int, float)):
                cell.number_format = "0.0%"
            elif header in MONEY_COLS and isinstance(cell.value, (int, float)):
                cell.number_format = "$#,##0.000"

    for c_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = COL_WIDTHS.get(header, 14)

    ws.freeze_panes = "A2"

    if last_row > 1:
        table_ref = f"A1:{get_column_letter(len(HEADERS))}{last_row}"
        table_name = sheet_name.replace(" ", "_") + "_Table"
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=False,
            showFirstColumn=False, showLastColumn=False, showColumnStripes=False,
        )
        ws.add_table(table)

    ws.row_dimensions[1].height = 30
    return ws


def write_legend_sheet(wb, run_timestamp):
    ws = wb.create_sheet("Read Me - Legend", 0)
    ws.column_dimensions["A"].width = 110
    lines = [
        ("ASX Strong Buy Stock Screener", True),
        (f"Report generated: {run_timestamp}", False),
        ("", False),
        ("HOW TO READ THIS REPORT", True),
        ("- Rows highlighted green have 'Strong Buy (Analysts)' = Yes.", False),
        ("- Click the filter arrows on any column header to sort/filter (e.g. filter Strong Buy = Yes only).", False),
        ("- 'ASX 200' tab covers the ASX 200 watchlist in tickers_asx200.csv.", False),
        ("- 'Penny Stocks' tab covers the speculative watchlist in tickers_penny_stocks.csv.", False),
        ("", False),
        ("DATA SOURCE & LIMITATIONS - PLEASE READ", True),
        ("- Source: Yahoo Finance (via the yfinance library). Free source, no paid data feed.", False),
        ("- Analyst coverage on Yahoo Finance for ASX stocks is inconsistent, especially for small/mid caps and", False),
        ("  almost all penny stocks. 'No data' means Yahoo does not publish that field for this stock.", False),
        ("- Forecast Growth (3/6/12 month): Yahoo (and free sources generally) only publish ONE analyst mean", False),
        ("  target price, which is conventionally a ~12-month horizon. The 3-month and 6-month figures in this", False),
        ("  report are a straight-line mathematical interpolation of that single 12-month figure, NOT independent", False),
        ("  3-month or 6-month analyst forecasts. Treat them as illustrative only.", False),
        ("- 'Strong Buy (Analysts)' = Yes when Yahoo's aggregated analyst recommendation score is 2.0 or better", False),
        ("  (scale: 1.0 = Strong Buy, 2.0 = Buy, 3.0 = Hold, 4.0 = Sell, 5.0 = Strong Sell). This is Yahoo's", False),
        ("  aggregated view of covering analysts, not a guarantee and not financial advice.", False),
        ("- Revenue / Net Profit (5yr): free yfinance access typically exposes only the last 4 annual reporting", False),
        ("  periods, not 5. Each figure is labelled with its actual reporting year.", False),
        ("- Notes - Business Status: this is the company's own static business description (Yahoo's", False),
        ("  'longBusinessSummary'), not a live or forward-looking analyst outlook. No free source publishes a", False),
        ("  structured 'current and future business status' commentary - a genuine forward outlook would need a", False),
        ("  paid research subscription (e.g. Morningstar, Refinitiv) or manual research per stock.", False),
        ("", False),
        ("THIS IS NOT FINANCIAL ADVICE. Verify all figures independently before making investment decisions.", True),
        ("", False),
        ("MAINTAINING THE STOCK LISTS", True),
        ("- Edit tickers_asx200.csv and tickers_penny_stocks.csv to add/remove stocks (one ASX code per line,", False),
        ("  no .AX suffix, no exchange prefix).", False),
        ("- The ASX 200 index is rebalanced quarterly (Mar/Jun/Sep/Dec) by S&P - review this list each quarter.", False),
        ("- There is no official 'penny stock' index; this list is a manually curated watchlist - edit freely.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(name="Arial", size=12 if bold and i == 1 else 10, bold=bold)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def build_report():
    run_ts = dt.datetime.now().strftime("%Y-%m-%d %H:%M %Z") or dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    log.info("Loading ticker lists...")
    asx200_tickers = load_tickers(HERE / "tickers_asx200.csv")
    penny_tickers = load_tickers(HERE / "tickers_penny_stocks.csv")

    log.info(f"ASX 200 universe: {len(asx200_tickers)} tickers")
    log.info(f"Penny Stocks universe: {len(penny_tickers)} tickers")

    asx200_df = fetch_universe(asx200_tickers, "ASX 200")
    penny_df = fetch_universe(penny_tickers, "Penny Stocks")

    wb = Workbook()
    wb.remove(wb.active)  # drop default blank sheet
    write_legend_sheet(wb, run_ts)
    write_sheet(wb, "ASX 200", asx200_df)
    write_sheet(wb, "Penny Stocks", penny_df)

    date_str = dt.date.today().isoformat()
    out_path = OUTPUT_DIR / f"ASX_Strong_Buy_Screener_{date_str}.xlsx"
    wb.save(out_path)
    log.info(f"Report written to {out_path}")
    return out_path


if __name__ == "__main__":
    build_report()
