"""
ASX 200 Strong Buy Screener — daily automated run.

What this does:
  1. Reads your ticker list from tickers.csv
  2. Pulls current price, 52-week high/low, and revenue figures for each
     ticker from Yahoo Finance (via the free `yfinance` library).
  3. Computes the same three signals as the manual workbook:
       - trading at/near its 52-week low
       - revenue growing year-on-year
       - a forward growth indicator (see CAVEAT below)
  4. Writes a two-sheet .xlsx (Data Input + Screener), identical layout
     to the manual version, plus two blank manual columns so you can
     jot down a Market Index or TradingView price for a quick 3-way
     cross-check on any flagged stock.
  5. Emails the workbook as an attachment.

CAVEAT — read this:
  Yahoo's free `revenueGrowth` field is a trailing (already-happened)
  year-on-year growth figure, not a true forward analyst forecast.
  Genuine forward revenue/earnings forecasts aren't available through
  any free, automatable source — Market Index and TradingView have
  analyst consensus figures but no free API, and scraping their pages
  on a schedule would breach their terms of use. So the "Growth
  Outlook OK?" column here is a same-metric proxy, clearly labelled.
  For a real forward view, check the manual comparison columns
  yourself for any stock the screener flags.

  This is a research shortlist tool, not financial advice.
"""

import os
import csv
import smtplib
from datetime import date
from email.message import EmailMessage

import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Adjustable screener settings
# ---------------------------------------------------------------------------
NEAR_LOW_THRESHOLD = 0.05        # flag if price is within 5% of the 52-week low
MIN_REVENUE_GROWTH = 0.00        # flag if YoY revenue growth >= 0%
MIN_FORECAST_GROWTH = 0.05       # flag if the growth proxy >= 5%
MIN_SIGNALS_FOR_STRONG_BUY = 3   # how many of the 3 signals must be met (1-3)

TICKERS_FILE = "tickers.csv"
OUTPUT_FILE = f"ASX200_Strong_Buy_Screener_{date.today().isoformat()}.xlsx"

FONT_NAME = "Arial"


def load_tickers(path):
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows.append((row["Ticker"].strip(), row["Company Name"].strip()))
    return rows


def fetch_stock_data(ticker):
    """Pull price/52-week/revenue data for one ASX ticker from Yahoo Finance.
    Returns a dict; any field that can't be retrieved is left as None so the
    workbook formulas degrade gracefully rather than erroring."""
    yf_ticker = yf.Ticker(f"{ticker}.AX")
    data = {
        "price": None, "year_high": None, "year_low": None,
        "latest_revenue": None, "prior_revenue": None,
        "growth_proxy": None,
    }

    try:
        hist = yf_ticker.history(period="1y")
        if not hist.empty:
            data["price"] = round(float(hist["Close"].iloc[-1]), 3)
            data["year_high"] = round(float(hist["High"].max()), 3)
            data["year_low"] = round(float(hist["Low"].min()), 3)
    except Exception as e:
        print(f"  [{ticker}] price history failed: {e}")

    try:
        fin = yf_ticker.financials  # annual income statement, most recent columns first
        if fin is not None and "Total Revenue" in fin.index and fin.shape[1] >= 2:
            data["latest_revenue"] = round(float(fin.loc["Total Revenue"].iloc[0]) / 1e6, 1)  # $m
            data["prior_revenue"] = round(float(fin.loc["Total Revenue"].iloc[1]) / 1e6, 1)
    except Exception as e:
        print(f"  [{ticker}] financials failed: {e}")

    try:
        info = yf_ticker.info
        g = info.get("revenueGrowth")
        if g is not None:
            data["growth_proxy"] = round(float(g), 4)
    except Exception as e:
        print(f"  [{ticker}] info/growth failed: {e}")

    return data


def build_workbook(rows_data):
    wb = openpyxl.Workbook()

    BLUE_BOLD = Font(name=FONT_NAME, color="0000FF", bold=True)
    BLACK = Font(name=FONT_NAME, color="000000")
    WHITE_BOLD = Font(name=FONT_NAME, color="FFFFFF", bold=True)
    TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="1F3864")
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    SETTINGS_FILL = PatternFill("solid", fgColor="FFF2CC")
    MANUAL_FILL = PatternFill("solid", fgColor="FCE4D6")
    STRONGBUY_FILL = PatternFill("solid", fgColor="C6EFCE")
    STRONGBUY_FONT = Font(name=FONT_NAME, color="006100", bold=True)
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def style_header(ws, row, last_col):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = WHITE_BOLD
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

    # ------------------------------------------------------------------
    # Data Input sheet
    # ------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Data Input"
    ws1.sheet_view.showGridLines = False

    headers = ["Ticker", "Company Name", "Current Price ($)", "52-Week High ($)", "52-Week Low ($)",
               "Latest FY Revenue ($m)", "Prior FY Revenue ($m)", "Yahoo Growth Proxy (trailing YoY %)",
               "Market Index Price (manual check)", "TradingView Price (manual check)", "Pulled On"]
    widths = [10, 30, 16, 16, 16, 18, 18, 22, 22, 22, 14]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws1.cell(row=1, column=1, value="ASX 200 Strong Buy Screener — Daily Data Pull (Yahoo Finance)").font = TITLE_FONT
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    note = ws1.cell(row=2, column=1,
                     value="Auto-pulled from Yahoo Finance. The two orange columns are for you to fill in "
                           "manually from Market Index / TradingView as a cross-check on flagged stocks — "
                           "these aren't pulled automatically (no free API on either site).")
    note.font = Font(name=FONT_NAME, italic=True, size=10, color="595959")
    note.alignment = Alignment(wrap_text=True)
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    ws1.row_dimensions[2].height = 30

    header_row = 4
    for i, h in enumerate(headers, start=1):
        ws1.cell(row=header_row, column=i, value=h)
    style_header(ws1, header_row, len(headers))
    ws1.row_dimensions[header_row].height = 34

    today_str = date.today().strftime("%d/%m/%Y")
    first_data_row = header_row + 1
    for offset, (ticker, name, d) in enumerate(rows_data):
        row = first_data_row + offset
        ws1.cell(row=row, column=1, value=ticker)
        ws1.cell(row=row, column=2, value=name)
        ws1.cell(row=row, column=3, value=d["price"])
        ws1.cell(row=row, column=4, value=d["year_high"])
        ws1.cell(row=row, column=5, value=d["year_low"])
        ws1.cell(row=row, column=6, value=d["latest_revenue"])
        ws1.cell(row=row, column=7, value=d["prior_revenue"])
        ws1.cell(row=row, column=8, value=d["growth_proxy"])
        ws1.cell(row=row, column=9, value=None)
        ws1.cell(row=row, column=10, value=None)
        ws1.cell(row=row, column=11, value=today_str)
        for c in range(1, 12):
            cell = ws1.cell(row=row, column=c)
            cell.border = BORDER
            cell.font = BLACK
            if c in (3, 4, 5, 9, 10):
                cell.number_format = '$#,##0.00'
            if c in (6, 7):
                cell.number_format = '$#,##0'
            if c == 8:
                cell.number_format = '0.0%'
            if c in (9, 10):
                cell.fill = MANUAL_FILL

    last_data_row = first_data_row + len(rows_data) - 1
    ws1.freeze_panes = "A5"

    # ------------------------------------------------------------------
    # Screener sheet
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Screener")
    ws2.sheet_view.showGridLines = False
    sc_widths = [10, 28, 14, 14, 16, 14, 16, 14, 18, 14, 12, 20]
    for i, w in enumerate(sc_widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.cell(row=1, column=1, value="ASX 200 Strong Buy Screener — Results").font = TITLE_FONT
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    ws2.cell(row=3, column=1, value="Settings (edit yellow cells)").font = Font(name=FONT_NAME, bold=True, size=12, color="1F3864")
    ws2.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)

    settings = [
        ("Near 52-Week-Low threshold (max % above the low)", NEAR_LOW_THRESHOLD, '0.0%'),
        ("Minimum revenue growth required (YoY)", MIN_REVENUE_GROWTH, '0.0%'),
        ("Minimum growth-proxy required", MIN_FORECAST_GROWTH, '0.0%'),
        ("Minimum signals met to flag STRONG BUY (1-3)", MIN_SIGNALS_FOR_STRONG_BUY, '0'),
    ]
    for i, (label, val, fmt) in enumerate(settings):
        row = 4 + i
        lbl = ws2.cell(row=row, column=1, value=label)
        lbl.font = BLACK
        lbl.alignment = Alignment(wrap_text=True)
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        val_cell = ws2.cell(row=row, column=4, value=val)
        val_cell.fill = SETTINGS_FILL
        val_cell.font = BLUE_BOLD
        val_cell.number_format = fmt
        val_cell.border = BORDER
        val_cell.alignment = Alignment(horizontal="center")
        ws2.row_dimensions[row].height = 28

    header_row2 = 9
    sc_headers = ["Ticker", "Company Name", "Current Price ($)", "52-Week Low ($)", "% Above 52W Low",
                  "Near 52W Low?", "Revenue Growth %", "Revenue Growing?", "Growth Proxy %",
                  "Growth Outlook OK?", "Signals Met", "Strong Buy Signal"]
    for i, h in enumerate(sc_headers, start=1):
        ws2.cell(row=header_row2, column=i, value=h)
    style_header(ws2, header_row2, len(sc_headers))
    ws2.row_dimensions[header_row2].height = 34

    n_rows = len(rows_data)
    di = "'Data Input'!"
    for offset in range(n_rows):
        src_row = first_data_row + offset
        row = header_row2 + 1 + offset
        ws2.cell(row=row, column=1, value=f"={di}A{src_row}")
        ws2.cell(row=row, column=2, value=f"={di}B{src_row}")
        ws2.cell(row=row, column=3, value=f"={di}C{src_row}")
        ws2.cell(row=row, column=4, value=f"={di}E{src_row}")
        ws2.cell(row=row, column=5, value=f"=IF(OR(D{row}=\"\",D{row}=0),\"\",(C{row}-D{row})/D{row})")
        ws2.cell(row=row, column=6, value=f"=IF(E{row}=\"\",\"\",IF(E{row}<=$D$4,\"YES\",\"\"))")
        ws2.cell(row=row, column=7,
                 value=f"=IF(OR({di}G{src_row}=\"\",{di}G{src_row}=0),\"\",({di}F{src_row}-{di}G{src_row})/{di}G{src_row})")
        ws2.cell(row=row, column=8, value=f"=IF(G{row}=\"\",\"\",IF(G{row}>=$D$5,\"YES\",\"\"))")
        ws2.cell(row=row, column=9, value=f"=IF({di}H{src_row}=\"\",\"\",{di}H{src_row})")
        ws2.cell(row=row, column=10, value=f"=IF(I{row}=\"\",\"\",IF(I{row}>=$D$6,\"YES\",\"\"))")
        ws2.cell(row=row, column=11, value=f"=(F{row}=\"YES\")+(H{row}=\"YES\")+(J{row}=\"YES\")")
        ws2.cell(row=row, column=12, value=f"=IF(A{row}=\"\",\"\",IF(K{row}>=$D$7,\"STRONG BUY\",\"\"))")
        for c in range(1, 13):
            cell = ws2.cell(row=row, column=c)
            cell.border = BORDER
            cell.font = BLACK
            if c in (3, 4):
                cell.number_format = '$#,##0.00'
            if c in (5, 7, 9):
                cell.number_format = '0.0%'
            if c in (6, 8, 10, 11, 12):
                cell.alignment = Alignment(horizontal="center")

    last_row = header_row2 + n_rows
    rng = f"A{header_row2+1}:L{last_row}"
    ws2.conditional_formatting.add(
        rng, FormulaRule(formula=[f"$L{header_row2+1}=\"STRONG BUY\""], fill=STRONGBUY_FILL, font=STRONGBUY_FONT)
    )
    ws2.freeze_panes = f"A{header_row2+1}"

    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


def send_email(attachment_path):
    gmail_address = os.environ["GMAIL_ADDRESS"]
    gmail_app_password = os.environ["GMAIL_APP_PASSWORD"]
    recipient = os.environ["RECIPIENT_EMAIL"]

    msg = EmailMessage()
    msg["Subject"] = f"ASX 200 Strong Buy Screener — {date.today().strftime('%d %b %Y')}"
    msg["From"] = gmail_address
    msg["To"] = recipient
    msg.set_content(
        "Attached is today's ASX 200 screener run.\n\n"
        "Data source: Yahoo Finance (automated). Market Index / TradingView columns "
        "on the Data Input tab are for your own manual cross-check.\n\n"
        "This is a research shortlist tool, not financial advice.\n"
    )

    with open(attachment_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(attachment_path),
        )

    with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
        smtp.starttls()
        smtp.login(gmail_address, gmail_app_password)
        smtp.send_message(msg)


def main():
    tickers = load_tickers(TICKERS_FILE)
    print(f"Pulling data for {len(tickers)} tickers from Yahoo Finance...")
    rows_data = []
    for ticker, name in tickers:
        print(f" - {ticker}")
        d = fetch_stock_data(ticker)
        rows_data.append((ticker, name, d))

    path = build_workbook(rows_data)
    print(f"Workbook written: {path}")

    send_email(path)
    print("Email sent.")


if __name__ == "__main__":
    main()
